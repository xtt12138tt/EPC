"""
Excel 模板解析器
读取 报销提报输入模板_v1.xlsx 并结构化输出
"""
import re
from pathlib import Path
import openpyxl


def read_kv_sheet(ws, key_col=1, val_col=2, skip_rows=1):
    """读取 KV 格式 Sheet（字段名 | 值 两列）"""
    data = {}
    for row in ws.iter_rows(min_row=1 + skip_rows, values_only=True):
        k = row[key_col - 1] if len(row) >= key_col else None
        v = row[val_col - 1] if len(row) >= val_col else None
        if k and str(k).strip() and not str(k).startswith("<"):
            data[str(k).strip()] = v
    return data


def read_table_sheet(ws, header_row=1, data_start=None, skip_hint_row=True):
    """
    读取表格格式 Sheet（第一行为表头，后续为数据行）
    skip_hint_row: 跳过含 '<删除本行说明>' 的提示行
    """
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value).strip() if cell.value else "")

    rows = []
    start = (data_start or header_row + 1)
    # 如果有总样本量行在 header 前，先单独处理
    for row in ws.iter_rows(min_row=start, values_only=True):
        if not any(row):
            continue
        first = str(row[0] or "")
        if skip_hint_row and "<删除本行说明>" in first:
            continue
        # 跳过参考说明块（以工作类型↔开头）
        if "↔" in first or first.startswith("工作类型") and "↔" in str(row[1] or ""):
            break
        record = {}
        for i, h in enumerate(headers):
            if h:
                record[h] = row[i] if i < len(row) else None
        if any(v is not None and str(v).strip() for v in record.values()):
            rows.append(record)
    return rows


def parse_excel(path: str) -> dict:
    """
    解析完整模板文件，返回结构化数据
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}

    # ── 01_项目概览 ──
    ws01 = wb["01_项目概览"]
    overview = read_kv_sheet(ws01, skip_rows=1)
    result["overview"] = {
        "project_id": _str(overview.get("项目号")),
        "project_name": _str(overview.get("项目名称")),
        "fanbao_series": _str(overview.get("发包序列")),
        "fanbao_type": _str(overview.get("发包类型")),
        "cost_center": _str(overview.get("成本归属")),
        "expense_note_manual": _str(overview.get("费用说明")),
        "screenshot_path": _str(overview.get("产品确认截图路径")),
        "reimbursement_types_raw": _str(overview.get("报销明细")),
    }
    # 解析多选报销明细
    raw = result["overview"]["reimbursement_types_raw"] or ""
    result["overview"]["reimbursement_types"] = [
        t.strip() for t in re.split(r"[、,，]", raw) if t.strip()
    ]

    # ── 02_费用说明_变量 ──
    ws02 = wb["02_费用说明_变量"]
    fee_vars = read_kv_sheet(ws02, skip_rows=1)
    result["fee_vars"] = {k: _str(v) for k, v in fee_vars.items()}

    # ── 03_礼金_小众特殊 ──
    ws03 = wb["03_礼金_小众特殊"]
    result["gift_special"] = read_table_sheet(ws03)

    # ── 04_礼金_常规 ──
    ws04 = wb["04_礼金_常规"]
    # 总样本量在 B1
    total_sample = ws04.cell(row=1, column=2).value
    result["gift_common"] = {
        "total_sample": _num(total_sample),
        "rows": read_table_sheet(ws04, header_row=3, data_start=5),
    }

    # ── 05_问卷调研 ──
    ws05 = wb["05_问卷调研"]
    result["questionnaire"] = read_table_sheet(ws05)

    # ── 06_国内兼职 ──
    ws06 = wb["06_国内兼职"]
    result["parttime"] = read_table_sheet(ws06)

    # ── 07_收款人明细 ──
    ws07 = wb["07_收款人明细"]
    # 跳过示例行（灰色，第2/3行），数据从第4行开始
    payees_raw = read_table_sheet(ws07, header_row=1, data_start=4)
    result["payees"] = [
        p for p in payees_raw
        if p.get("玩家姓名") and str(p.get("玩家姓名", "")).strip()
           and str(p.get("玩家姓名", "")).strip() not in ("张三", "Mark")
    ]

    # ── 08_拆单备注 ──
    ws08 = wb["08_拆单备注"]
    split_data = read_kv_sheet(ws08, skip_rows=1)
    result["split_note"] = {
        "enabled": str(split_data.get("是否拆单（Y/N）", "N")).upper() == "Y",
        "note": _str(split_data.get("拆单备注文本")),
    }

    wb.close()
    return result


def build_expense_note(data: dict) -> str:
    """
    根据 fee_vars 自动拼装统一格式的费用说明文本
    如果用户在 overview.expense_note_manual 已手写，直接返回手写内容
    """
    if data["overview"].get("expense_note_manual"):
        return data["overview"]["expense_note_manual"]

    v = data.get("fee_vars", {})

    def _val(key):
        return v.get(key) or ""

    product = _val("产品名称")
    project = _val("项目名称")
    total = _val("EPC报销金额") or _val("总费用")
    epc_amount = _val("EPC报销金额")
    other_method = _val("其他报销方式")
    other_amount = _val("已报销金额")

    header = f"{product}执行了{project}项目，共产生费用{total}元，全部走EPC公账报销。"
    if other_method and other_amount and epc_amount:
        header += f"（已通过{other_method}报销了{other_amount}元，本次EPC报销{epc_amount}元。）"
    lines = [header, "详细报销内容为："]

    # 玩家礼金
    gift_amount = _val("玩家礼金金额")
    gift_count = _val("玩家数量")
    gift_price = _val("礼金单价")
    gift_extra = _val("玩家礼金补充说明")
    if gift_amount:
        s = f"【玩家礼金】{gift_amount}元"
        if gift_count and gift_price:
            s += f"，共{gift_count}名玩家参与测试，样本单价{gift_price}元/人"
        if gift_extra:
            s += f"（{gift_extra}）"
        s += "。"
        lines.append(s)

    # 兼职费用
    pt_amount = _val("兼职费用金额")
    pt_count = _val("兼职数量描述")
    pt_price = _val("兼职单价")
    pt_extra = _val("兼职补充说明")
    if pt_amount:
        s = f"【兼职费用】{pt_amount}元"
        if pt_count and pt_price:
            s += f"，明细为：测试执行{pt_count}×{pt_price}={pt_amount}元"
        if pt_extra:
            s += f"（{pt_extra}）"
        s += "。"
        lines.append(s)

    # 餐饮、交通、场地统一归入其他费用
    other_details = []
    other_total = 0.0
    meal_amount = _val("餐饮金额")
    meal_count = _val("餐饮人数")
    meal_price = _val("餐补单价")
    if meal_amount:
        other_total += _num(meal_amount) or 0
        if meal_count and meal_price:
            other_details.append(f"餐饮费共计{meal_count}人次，单价{meal_price}元/人次，共计{meal_count}×{meal_price}={meal_amount}元")
        else:
            other_details.append(f"餐饮费{meal_amount}元")

    trans_amount = _val("交通金额")
    trans_count = _val("交通人数")
    trans_price = _val("交通单价")
    trans_reason = _val("交通原因")
    if trans_amount:
        other_total += _num(trans_amount) or 0
        if trans_count and trans_price:
            other_details.append(f"交通费共计{trans_count}人次，单价{trans_price}元/人次，共计{trans_count}×{trans_price}={trans_amount}元")
        else:
            other_details.append(f"交通费{trans_amount}元")

    venue_amount = _val("场地金额")
    venue_count = _val("场地场次")
    venue_price = _val("场地单价")
    if venue_amount:
        other_total += _num(venue_amount) or 0
        if venue_count and venue_price:
            other_details.append(f"场地/设备租赁{venue_count}场×{venue_price}={venue_amount}元")
        else:
            other_details.append(f"场地/设备租赁{venue_amount}元")

    if other_details:
        other_total_text = str(int(other_total)) if other_total.is_integer() else str(other_total)
        lines.append(f"【其他费用】{other_total_text}元，明细为：{'、'.join(other_details)}。")

    return "\n\n".join(lines).strip()


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _num(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None
